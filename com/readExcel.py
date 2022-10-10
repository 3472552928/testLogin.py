import os
import openpyxl


class ReadExcel():
    def __init__(self,filename,sheetname):
        path = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
        self.filename=os.path.join(path,filename)
        self.sheetname=sheetname
        execl = openpyxl.load_workbook(self.filename)
        self.sheet = execl[self.sheetname]

    # # def read_excel(self):
    #     print(self.sheet.max_row,self.sheet.max_column)



    def readall(self,endrow,endcol,startcol=1,startrow=2):
        res = []
        if endrow >self.sheet.max_row:
            endrow = self.sheet.max_row
        if endcol >self.sheet.max_column:
            endcol = self.sheet.max_column
        for i in range(startrow,endrow+1):
            lis = []
            for line in range(startcol,endcol+1):
                lis.append(self.sheet.cell(i,line).value)
            res.append(lis)
        print(res)


if __name__ == '__main__':
    filename = "conf\data1.xlsx"
    re = ReadExcel(filename,"Sheet1")
    re.readall(startrow=2,endrow=7,startcol=1,endcol=4)
